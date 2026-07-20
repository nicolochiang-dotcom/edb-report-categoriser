"""Run the categoriser over every report in reports_in/ and append the
results to ONE persistent Excel workbook in grid format.

Layout:
  Column A : Company Name / Report      (the file name)
  Column B : Home Country
  Column C : Industry
  Column D : Engagement Type
  Column E : Date (DD/MM/YYYY)
  Columns F..AA : one column per category (ids 1..22)

Each report is one row. Cells hold the passage(s) for that report+category,
prefixed with confidence. Re-running a report already in the workbook
overwrites that report's row rather than duplicating it.
"""

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from categories import CATEGORIES
from classifier import classify_report
from ingest import SUPPORTED_EXTENSIONS, extract_text

INPUT_DIR = Path("reports_in")
OUTPUT_DIR = Path("output")
MASTER_FILE = OUTPUT_DIR / "categorised_reports_master.xlsx"

CATEGORY_NAMES = {c["id"]: c["name"] for c in CATEGORIES}
CATEGORY_IDS = [c["id"] for c in CATEGORIES]  # [1, 2, ... 22]

# Metadata columns: (excel header, key returned by the classifier)
METADATA_COLUMNS = [
    ("Home Country", "home_country"),
    ("Industry", "industry"),
    ("Engagement Type", "engagement_type"),
    ("Date", "report_date"),
]

# Column A = report name, then metadata, then the categories.
FIRST_META_COL = 2                                    # column B
FIRST_CAT_COL = FIRST_META_COL + len(METADATA_COLUMNS)  # column F
LAST_COL = FIRST_CAT_COL + len(CATEGORY_IDS) - 1        # column AA


def col_for_category(cat_id: int) -> int:
    return FIRST_CAT_COL + CATEGORY_IDS.index(cat_id)


def build_header(ws) -> None:
    """Write row 1: report name, metadata columns, then the 22 categories."""
    meta_fill = PatternFill("solid", fgColor="2F5597")
    cat_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    wrap_top = Alignment(wrap_text=True, vertical="top")

    ws.cell(row=1, column=1, value="Company Name / Report")

    for i, (header, _key) in enumerate(METADATA_COLUMNS):
        ws.cell(row=1, column=FIRST_META_COL + i, value=header)

    for cat_id in CATEGORY_IDS:
        ws.cell(
            row=1,
            column=col_for_category(cat_id),
            value=f"Category {cat_id}: {CATEGORY_NAMES[cat_id]}",
        )

    for col in range(1, LAST_COL + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = meta_fill if col < FIRST_CAT_COL else cat_fill
        cell.font = header_font
        cell.alignment = wrap_top

    # Column widths.
    ws.column_dimensions["A"].width = 26
    for i in range(len(METADATA_COLUMNS)):
        letter = get_column_letter(FIRST_META_COL + i)
        ws.column_dimensions[letter].width = 18
    for cat_id in CATEGORY_IDS:
        letter = get_column_letter(col_for_category(cat_id))
        ws.column_dimensions[letter].width = 40

    # Freeze the header row and everything left of the first category column.
    ws.freeze_panes = f"{get_column_letter(FIRST_CAT_COL)}2"


def load_or_create_workbook():
    """Open the master workbook if it exists, else create it with headers."""
    if MASTER_FILE.exists():
        wb = load_workbook(MASTER_FILE)
        return wb, wb.active

    wb = Workbook()
    ws = wb.active
    ws.title = "Extractions"
    build_header(ws)
    return wb, ws


def find_report_row(ws, report_name: str):
    """Return the existing row number for this report, or None."""
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == report_name:
            return row
    return None


def format_cell(passages: list[dict]) -> str:
    """Combine one report's passages for a single category into cell text."""
    return "\n\n".join(
        f"Confidence={p['confidence'].capitalize()}. {p['passage']}"
        for p in passages
    )


def write_report_row(ws, report_name, metadata, extractions) -> None:
    """Write (or overwrite) one report's row."""
    by_cat: dict[int, list[dict]] = {}
    for item in extractions:
        for cat_id in item["categories"]:
            by_cat.setdefault(cat_id, []).append(item)

    row = find_report_row(ws, report_name)
    if row is None:
        row = max(ws.max_row + 1, 2)

    wrap_top = Alignment(wrap_text=True, vertical="top")

    ws.cell(row=row, column=1, value=report_name).alignment = wrap_top

    for i, (_header, key) in enumerate(METADATA_COLUMNS):
        cell = ws.cell(row=row, column=FIRST_META_COL + i)
        cell.value = metadata.get(key) or None
        cell.alignment = wrap_top

    for cat_id in CATEGORY_IDS:
        cell = ws.cell(row=row, column=col_for_category(cat_id))
        cell.value = format_cell(by_cat[cat_id]) if cat_id in by_cat else None
        cell.alignment = wrap_top


def main() -> None:
    OUTPUT_DIR.mkdir(exist_ok=True)
    files = sorted(
        p for p in INPUT_DIR.iterdir()
        if p.suffix.lower() in SUPPORTED_EXTENSIONS
    )
    if not files:
        print(f"No reports found in {INPUT_DIR}/ - drop some files in and rerun.")
        return

    wb, ws = load_or_create_workbook()
    processed, failures = 0, []

    for i, path in enumerate(files, start=1):
        print(f"[{i}/{len(files)}] {path.name}")
        try:
            text = extract_text(path)
            if len(text.strip()) < 50:
                raise ValueError(
                    "almost no text extracted (scanned PDF or empty file?)"
                )
            metadata, extractions = classify_report(text)
        except Exception as err:
            print(f"    FAILED: {err}")
            failures.append(path.name)
            continue

        write_report_row(ws, path.name, metadata, extractions)
        processed += 1
        print(
            f"    {metadata.get('home_country') or '?'} | "
            f"{metadata.get('industry') or '?'} | "
            f"{metadata.get('engagement_type') or '?'} | "
            f"{metadata.get('report_date') or '?'}"
        )
        print(f"    {len(extractions)} passage(s) extracted.")

    wb.save(MASTER_FILE)
    print(f"\nDone: {processed}/{len(files)} report(s) written.")
    if failures:
        print(f"{len(failures)} file(s) failed: {', '.join(failures)}")
    print(f"Master workbook: {MASTER_FILE}")


if __name__ == "__main__":
    main()
